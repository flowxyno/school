import openpyxl

# Path to the directory containing the data
directory_path = r'S:\Testing_Ground\SampleData'

# Load the workbook as wb
wb = openpyxl.load_workbook(directory_path + r'\MTG_Card_Sample.xlsx')

# Get the active sheet
sheet = wb.active

# Create a new workbook to save requested data in
wb_rare = openpyxl.Workbook()

# Get the active sheet in the new workbook
sheet_rare = wb_rare.active

# Copy header row for column names
for col in range(1, sheet.max_column + 1):
    sheet_rare.cell(row=1, column=col).value = sheet.cell(row=1, column=col).value

# Copy data for rare cards
row_rare = 2
for row in range(2, sheet.max_row + 1):
    rarity = sheet.cell(row=row, column=2).value
    if rarity == 'Rare':
        for col in range(1, sheet.max_column + 1):
            sheet_rare.cell(row=row_rare, column=col).value = sheet.cell(row=row, column=col).value

        row_rare += 1

# Save the new workbook
wb_rare.save(directory_path + r'\MTG_Rare_Cards.xlsx')
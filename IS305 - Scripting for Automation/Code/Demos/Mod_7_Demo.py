import openpyxl

# Path to the directory containing the data
workdir = r'S:\Testing_Ground\SampleData'

# Load the workbook as wb
wb = openpyxl.load_workbook(workdir + r'\MTG_Card_Sample.xlsx')

# Get the active sheet
origSheet = wb.active

# Create a new workbook to save requested data in
rare_wb = openpyxl.Workbook()

# Get the active sheet in the new workbook
rareSheet = rare_wb.active

# Copy header row for column names
for col in range(1, origSheet.max_column +1):
    rareSheet.cell(row=1, column=col).value = origSheet.cell(row=1, column=col).value

# Copy data for rare cards
rareRow = 2
for row in range(2, origSheet.max_row +1):
    rarity = origSheet.cell(row=row, column=2).value
    if rarity == 'Rare':
        for col in range(1, origSheet.max_column +1):
            rareSheet.cell(row=rareRow, column=col).value = origSheet.cell(row=row, column=col).value
        rareRow += 1

# Save the new workbook
rare_wb.save(workdir + r'\MTG_Rare_Cards.xlsx')
